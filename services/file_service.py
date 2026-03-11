"""
BOM Compare - File Service
Handles reading and writing Excel (.xlsx, .xls) and CSV files.
"""
import csv
import logging
import os
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

import openpyxl
import xlrd

# Patch openpyxl CellStyle to handle 'xxid' attribute issue in some Excel files
try:
    from openpyxl.styles.cell_style import CellStyle
    _original_init = CellStyle.__init__
    def _patched_init(self, *args, **kwargs):
        # Remove 'xxid' if present, it's not a valid parameter
        kwargs.pop('xxid', None)
        _original_init(self, *args, **kwargs)
    CellStyle.__init__ = _patched_init
except Exception:
    pass  # If patch fails, continue without it


def allowed_file(filename: str, allowed_extensions: set) -> bool:
    """Check if file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def detect_csv_delimiter(file_path: str, sample_size: int = 5) -> str:
    """
    Auto-detect CSV delimiter by analyzing the first few lines.
    Supports: comma, semicolon, tab, pipe
    """
    delimiters = [',', ';', '\t', '|']
    
    # Try multiple encodings
    encodings = ['utf-8-sig', 'latin-1', 'cp1252']
    sample = []
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                sample = [f.readline() for _ in range(sample_size)]
            break
        except UnicodeDecodeError:
            continue
    
    if not sample:
        # Fallback to latin-1 which never fails
        with open(file_path, 'r', encoding='latin-1') as f:
            sample = [f.readline() for _ in range(sample_size)]
    
    # Count occurrences of each delimiter
    counts = {}
    for delim in delimiters:
        counts[delim] = sum(line.count(delim) for line in sample)
    
    # Return the delimiter with highest count
    best_delim = max(counts, key=counts.get)
    return best_delim if counts[best_delim] > 0 else ','


def detect_header_row(file_path: str, delimiter: str = None, max_rows: int = 10) -> int:
    """
    Auto-detect the header row by finding the first row with multiple columns.
    This handles CSV files with metadata/title rows at the top.
    
    Args:
        file_path: Path to CSV file
        delimiter: Delimiter to use (auto-detected if None)
        max_rows: Maximum rows to check
    
    Returns:
        0-indexed row number where headers are located
    """
    if delimiter is None:
        delimiter = detect_csv_delimiter(file_path)
    
    # Try multiple encodings
    encodings = ['utf-8-sig', 'latin-1', 'cp1252']
    lines = []
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = [f.readline() for _ in range(max_rows)]
            break
        except UnicodeDecodeError:
            continue
    
    if not lines:
        with open(file_path, 'r', encoding='latin-1') as f:
            lines = [f.readline() for _ in range(max_rows)]
    
    # Find first row with multiple columns (likely the header)
    for i, line in enumerate(lines):
        parts = line.strip().split(delimiter)
        # A header row should have multiple columns
        if len(parts) >= 3:
            return i
    
    return 0  # Default to first row


def read_csv(file_path: str, header_row: int = 0) -> tuple[list[str], list[dict]]:
    """
    Read CSV file and return headers and rows as list of dicts.
    
    Args:
        file_path: Path to CSV file
        header_row: 0-indexed row number for headers (default: 0)
    
    Returns:
        Tuple of (headers list, rows as list of dicts)
    """
    delimiter = detect_csv_delimiter(file_path)
    
    # Try multiple encodings
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                # Skip rows before header
                for _ in range(header_row):
                    next(f)
                
                reader = csv.DictReader(f, delimiter=delimiter)
                raw_headers = reader.fieldnames or []
                # Filter out None headers and convert to valid strings
                headers = [str(h) if h is not None else f'Column_{i}' for i, h in enumerate(raw_headers)]
                
                # Read rows and convert None values to empty strings
                rows = []
                for row in reader:
                    clean_row = {}
                    for key, value in row.items():
                        # Handle None keys (empty column headers)
                        clean_key = str(key) if key is not None else ''
                        # Convert None values to empty strings
                        clean_value = str(value) if value is not None else ''
                        if clean_key:  # Only include if key is not empty
                            clean_row[clean_key] = clean_value
                    rows.append(clean_row)
            
            return headers, rows
        except UnicodeDecodeError:
            continue
    
    # If all encodings fail, use latin-1 which never fails (may have wrong characters)
    with open(file_path, 'r', encoding='latin-1', errors='replace') as f:
        for _ in range(header_row):
            next(f)
        
        reader = csv.DictReader(f, delimiter=delimiter)
        raw_headers = reader.fieldnames or []
        headers = [str(h) if h is not None else f'Column_{i}' for i, h in enumerate(raw_headers)]
        
        rows = []
        for row in reader:
            clean_row = {}
            for key, value in row.items():
                clean_key = str(key) if key is not None else ''
                clean_value = str(value) if value is not None else ''
                if clean_key:
                    clean_row[clean_key] = clean_value
            rows.append(clean_row)
    
    return headers, rows


def get_sheet_names(file_path: str) -> list[str]:
    """
    Get list of sheet names from an Excel file.
    
    Args:
        file_path: Path to Excel file (.xlsx or .xls)
    
    Returns:
        List of sheet names, or empty list for non-Excel files
    """
    ext = Path(file_path).suffix.lower()
    
    if ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            logger.warning(f"Could not read sheet names from xlsx: {e}")
            return []
    elif ext == '.xls':
        try:
            wb = xlrd.open_workbook(file_path)
            return wb.sheet_names()
        except Exception as e:
            logger.warning(f"Could not read sheet names from xls: {e}")
            return []
    else:
        return []  # CSV and other formats don't have sheets


def read_xlsx(file_path: str, header_row: int = 0, sheet_name: Optional[str] = None) -> tuple[list[str], list[dict]]:
    """
    Read .xlsx file using openpyxl.
    
    Args:
        file_path: Path to Excel file
        header_row: 0-indexed row number for headers (default: 0)
        sheet_name: Optional name of sheet to read (default: active sheet)
    
    Returns:
        Tuple of (headers list, rows as list of dicts)
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except TypeError as e:
        # Handle openpyxl compatibility issues with certain Excel files
        if 'xxid' in str(e):
            # Workaround: try loading without read_only mode
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception:
                raise ValueError(f"Unable to read Excel file. Try re-saving the file in Excel to fix formatting issues: {e}")
        else:
            raise
    
    # Select sheet by name or use active sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Collect rows, skipping fully empty rows to handle files with
    # oversized sheet dimensions (e.g. 1M+ rows where most are empty)
    rows_data = []
    consecutive_empty = 0
    max_empty_gap = 300  # Stop after this many consecutive empty rows
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            # If we skipped some empty rows within data, add them back
            for _ in range(consecutive_empty):
                rows_data.append(tuple(None for _ in range(len(row))))
            consecutive_empty = 0
            rows_data.append(row)
        else:
            consecutive_empty += 1
            if rows_data and consecutive_empty >= max_empty_gap:
                break
    wb.close()

    if not rows_data or len(rows_data) <= header_row:
        return [], []

    # Get headers from specified row
    headers = [str(cell) if cell is not None else f'Column_{i}'
               for i, cell in enumerate(rows_data[header_row])]

    # Get data rows (after header)
    result_rows = []
    for row in rows_data[header_row + 1:]:
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell) if cell is not None else ''
        result_rows.append(row_dict)

    return headers, result_rows


def read_xls(file_path: str, header_row: int = 0, sheet_name: Optional[str] = None) -> tuple[list[str], list[dict]]:
    """
    Read legacy .xls file using xlrd.
    
    Args:
        file_path: Path to Excel file
        header_row: 0-indexed row number for headers (default: 0)
        sheet_name: Optional name of sheet to read (default: first sheet)
    
    Returns:
        Tuple of (headers list, rows as list of dicts)
    """
    wb = xlrd.open_workbook(file_path)
    
    # Select sheet by name or use first sheet
    if sheet_name and sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)
    
    if ws.nrows <= header_row:
        return [], []
    
    # Get headers
    headers = [str(cell.value) if cell.value else f'Column_{i}' 
               for i, cell in enumerate(ws.row(header_row))]
    
    # Get data rows
    result_rows = []
    for row_idx in range(header_row + 1, ws.nrows):
        row = ws.row(row_idx)
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell.value) if cell.value else ''
        result_rows.append(row_dict)
    
    return headers, result_rows


def read_file(
    file_path: str, 
    header_row: int = 0,
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
    sheet_name: Optional[str] = None
) -> tuple[list[str], list[dict]]:
    """
    Read file based on extension. Supports .xlsx, .xls, .csv
    
    Args:
        file_path: Path to file
        header_row: 0-indexed row number for headers (default: 0)
        start_row: 0-indexed first data row to include (default: all rows)
        end_row: 0-indexed last data row to include (default: all rows)
        sheet_name: Optional sheet name for Excel files (default: active/first sheet)
    
    Returns:
        Tuple of (headers list, rows as list of dicts)
    
    Raises:
        ValueError: If file format is not supported
    """
    ext = Path(file_path).suffix.lower()
    
    if ext == '.xlsx':
        headers, rows = read_xlsx(file_path, header_row, sheet_name)
    elif ext == '.xls':
        headers, rows = read_xls(file_path, header_row, sheet_name)
    elif ext == '.csv':
        headers, rows = read_csv(file_path, header_row)
    else:
        raise ValueError(f"Unsupported file format: {ext}")
    
    # Apply row range filter if specified
    if start_row is not None or end_row is not None:
        start = start_row if start_row is not None else 0
        end = end_row if end_row is not None else len(rows)
        rows = rows[start:end + 1]  # +1 because end_row is inclusive
    
    return headers, rows


def get_file_preview(file_path: str, header_row: int = 0, max_rows: int = 5) -> dict:
    """
    Get a preview of the file for display in the UI.
    
    Returns:
        Dict with 'headers', 'rows' (limited), 'total_rows', 'file_type'
    """
    headers, rows = read_file(file_path, header_row)
    
    return {
        'headers': headers,
        'rows': rows[:max_rows],
        'total_rows': len(rows),
        'file_type': Path(file_path).suffix.lower()
    }


def write_comparison_excel(
    comparison_result: dict,
    output_path: str,
    bom1_name: str = "Faber BOM",
    bom2_name: str = "Client BOM",
    params: list = None,
    regex_params: list = None,
    main_bom: int = 1
) -> None:
    """
    Export comparison result to Excel with color formatting.
    
    Args:
        comparison_result: Result from compare_service.compare_boms()
        output_path: Path to save the Excel file
        bom1_name: Name for first BOM sheet/header
        bom2_name: Name for second BOM sheet/header
        params: Optional list of extracted parameters from AI analysis
        regex_params: Optional list of extracted parameters from regex extraction
        main_bom: Which BOM is the main reference (1 or 2)
    """
    from openpyxl.styles import PatternFill, Font
    
    # Define colors
    MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # LightGreen
    DIFF_FILL = PatternFill(start_color="DB7093", end_color="DB7093", fill_type="solid")   # PaleVioletRed
    ACCEPTABLE_FILL = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")  # Gold/Amber
    UNMATCHED_FILL = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # MistyRose
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark blue
    PARAM_HEADER_FILL = PatternFill(start_color="5C6BC0", end_color="5C6BC0", fill_type="solid")  # Indigo
    MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    columns = comparison_result.get('columns', [])
    rows = comparison_result.get('rows', [])
    
    # Add main BOM indicator text in header
    main_indicator = " ⭐ MAIN BOM"
    bom1_label = bom1_name + (main_indicator if main_bom == 1 else "")
    bom2_label = bom2_name + (main_indicator if main_bom == 2 else "")
    
    # Write headers for BOM1
    col_offset = 1
    for i, col in enumerate(columns):
        cell = ws.cell(row=1, column=col_offset + i, value=f"{bom1_label} - {col}" if i == 0 else col)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Gap columns
    gap_start = col_offset + len(columns)
    ws.cell(row=1, column=gap_start, value="")
    ws.cell(row=1, column=gap_start + 1, value="")
    
    # Write headers for BOM2
    col_offset_bom2 = gap_start + 2
    for i, col in enumerate(columns):
        cell = ws.cell(row=1, column=col_offset_bom2 + i, value=f"{bom2_label} - {col}" if i == 0 else col)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Parameters columns (if params or regex_params provided)
    param_cols = ['Value', 'Tolerance', 'Voltage', 'Power', 'Package']
    param_col_start = None
    has_params = params or regex_params
    
    if has_params:
        # Add gap
        gap_start_params = col_offset_bom2 + len(columns)
        ws.cell(row=1, column=gap_start_params, value="")
        
        # Add param headers
        param_col_start = gap_start_params + 1
        for i, pcol in enumerate(param_cols):
            cell = ws.cell(row=1, column=param_col_start + i, value=f"Param - {pcol}")
            cell.fill = PARAM_HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
    
    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        bom1 = row_data.get('bom1', {})
        bom2 = row_data.get('bom2', {})
        status = row_data.get('status', 'matched')
        cell_diffs = row_data.get('cell_diffs', {})
        
        # Write BOM1 cells
        for col_idx, col in enumerate(columns):
            cell = ws.cell(row=row_idx, column=1 + col_idx, value=bom1.get(col, ''))
            
            if status in ('only_bom1', 'only_bom2'):
                cell.fill = UNMATCHED_FILL
            elif cell_diffs.get(col) == 'match':
                cell.fill = MATCH_FILL
            elif cell_diffs.get(col) == 'diff':
                cell.fill = DIFF_FILL
        
        # Write BOM2 cells
        for col_idx, col in enumerate(columns):
            cell = ws.cell(row=row_idx, column=col_offset_bom2 + col_idx, value=bom2.get(col, ''))
            
            if status in ('only_bom1', 'only_bom2'):
                cell.fill = UNMATCHED_FILL
            elif cell_diffs.get(col) == 'match':
                cell.fill = MATCH_FILL
            elif cell_diffs.get(col) == 'diff':
                cell.fill = DIFF_FILL
        
        # Write params cells - prefer regex_params, fall back to AI params
        if param_col_start:
            idx = row_idx - 2
            
            # Try regex params first
            if regex_params and idx < len(regex_params):
                param_data = regex_params[idx]
                if param_data:
                    # Regex params structure: value1, value2, valueStatus, tolerance1, etc.
                    param_map = [
                        ('value1', 'value2', 'valueStatus'),
                        ('tolerance1', 'tolerance2', 'toleranceStatus'),
                        ('voltage1', 'voltage2', 'voltageStatus'),
                        ('power1', 'power2', 'powerStatus'),
                        ('package1', 'package2', 'packageStatus')
                    ]
                    
                    for i, (key1, key2, status_key) in enumerate(param_map):
                        v1 = param_data.get(key1, '') or ''
                        v2 = param_data.get(key2, '') or ''
                        
                        if v1 and v2:
                            value = f"{v1} | {v2}"
                        elif v1:
                            value = f"{v1} | -"
                        elif v2:
                            value = f"- | {v2}"
                        else:
                            value = ""
                        
                        cell = ws.cell(row=row_idx, column=param_col_start + i, value=value)
                        
                        # Color based on match status
                        match_status = param_data.get(status_key, '')
                        if match_status == 'match':
                            cell.fill = MATCH_FILL
                        elif match_status == 'acceptable':
                            cell.fill = ACCEPTABLE_FILL
                        elif match_status == 'diff':
                            cell.fill = DIFF_FILL
                        elif match_status == 'missing':
                            cell.fill = MISSING_FILL
                    
                    continue  # Skip AI params for this row
            
            # Fall back to AI params
            if params and idx < len(params):
                param_data = params[idx]
                if param_data and param_data.get('component_type'):
                    # This is an R/C component
                    p1 = param_data.get('params_bom1') or {}
                    p2 = param_data.get('params_bom2') or {}
                    matches = param_data.get('param_matches') or {}
                    
                    param_keys = ['value', 'tolerance', 'voltage', 'power', 'package']
                    for i, pkey in enumerate(param_keys):
                        # Format value: show both BOM values
                        v1 = p1.get(pkey, '') or ''
                        v2 = p2.get(pkey, '') or ''
                        
                        if v1 and v2:
                            value = f"{v1} | {v2}"
                        elif v1:
                            value = f"{v1} | -"
                        elif v2:
                            value = f"- | {v2}"
                        else:
                            value = ""
                        
                        cell = ws.cell(row=row_idx, column=param_col_start + i, value=value)
                        
                        # Color based on match status
                        match_status = matches.get(pkey, '')
                        if match_status == 'match':
                            cell.fill = MATCH_FILL
                        elif match_status == 'acceptable':
                            cell.fill = ACCEPTABLE_FILL
                        elif match_status == 'diff':
                            cell.fill = DIFF_FILL
                        elif match_status == 'missing':
                            cell.fill = MISSING_FILL
    
    wb.save(output_path)

