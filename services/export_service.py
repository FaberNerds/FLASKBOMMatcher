"""
BOM Matcher - Export Service
Generates Excel export with added FaberNr and MPNfree columns.
"""
import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font

logger = logging.getLogger(__name__)

# Color fills
MATCHED_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Green
SUGGESTION_FILL = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")  # Yellow
NO_MATCH_FILL = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")    # Red
MPNFREE_FILL = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")     # Gray
HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")      # Dark blue


def export_bom(
    headers: list[str],
    rows: list[dict],
    matches: Optional[dict],
    mpnfree: Optional[dict],
    selections: Optional[dict],
    output_path: str
) -> str:
    """
    Export BOM to Excel with added FaberNr and MPNfree columns.

    Args:
        headers: Original BOM column headers
        rows: Original BOM row dicts
        matches: Match results from match_service (keyed by row index as string)
        mpnfree: MPNfree assessments (keyed by row index as string)
        selections: User overrides (keyed by row index as string)
        output_path: Path to save the Excel file

    Returns:
        The output path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    matches = matches or {}
    mpnfree = mpnfree or {}
    selections = selections or {}

    # Write headers: original columns + FaberNr + MPNfree?
    all_headers = list(headers) + ['FaberNr', 'MPNfree?']
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)

    # Write data rows
    for row_idx, row in enumerate(rows):
        excel_row = row_idx + 2  # 1-indexed, skip header
        str_idx = str(row_idx)

        # Write original columns
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=excel_row, column=col_idx, value=row.get(header, ''))

        # Determine FaberNr value
        fabernr = ''
        confidence = 'none'

        # Check user selections first (overrides)
        sel = selections.get(str_idx, {})
        if sel.get('fabernr'):
            fabernr = sel['fabernr']
            confidence = 'high'
        elif str_idx in matches:
            match_data = matches[str_idx]
            auto = match_data.get('auto_selected')
            if auto:
                fabernr = auto.get('FaberNr', '')
                confidence = match_data.get('confidence', 'low')

        # Determine MPNfree value
        is_mpnfree = ''
        if sel.get('mpnfree') is not None:
            is_mpnfree = 'Yes' if sel['mpnfree'] else 'No'
        elif str_idx in mpnfree:
            mpnfree_data = mpnfree[str_idx]
            is_mpnfree = 'Yes' if mpnfree_data.get('mpnfree') else 'No'

        # Write FaberNr
        fabernr_col = len(headers) + 1
        fabernr_cell = ws.cell(row=excel_row, column=fabernr_col, value=fabernr)

        # Write MPNfree
        mpnfree_col = len(headers) + 2
        mpnfree_cell = ws.cell(row=excel_row, column=mpnfree_col, value=is_mpnfree)

        # Apply color coding
        if is_mpnfree == 'Yes':
            fabernr_cell.fill = MPNFREE_FILL
            mpnfree_cell.fill = MPNFREE_FILL
        elif confidence == 'high':
            fabernr_cell.fill = MATCHED_FILL
        elif confidence in ('medium', 'low') and fabernr:
            fabernr_cell.fill = SUGGESTION_FILL
        elif not fabernr:
            fabernr_cell.fill = NO_MATCH_FILL

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(output_path)
    logger.info(f"Exported BOM to {output_path}")
    return output_path
